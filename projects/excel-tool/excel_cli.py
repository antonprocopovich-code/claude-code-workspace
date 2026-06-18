#!/usr/bin/env python3
"""Excel CLI tool for Claude Code. Returns compact JSON to minimize context tokens."""

import argparse
import json
import sys
from pathlib import Path


def _wb(path: str, write: bool = False):
    import openpyxl
    p = Path(path)
    if write and not p.exists():
        wb = openpyxl.Workbook()
        wb.save(path)
    return openpyxl.load_workbook(path, data_only=True) if not write else openpyxl.load_workbook(path)


def _out(data):
    print(json.dumps(data, ensure_ascii=False, default=str))


def cmd_list_sheets(args):
    wb = _wb(args.file)
    _out({"sheets": wb.sheetnames})


def cmd_info(args):
    wb = _wb(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    _out({
        "sheet": ws.title,
        "dims": ws.dimensions,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "sheets": wb.sheetnames,
    })


def cmd_read(args):
    wb = _wb(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    if args.range:
        cells = ws[args.range]
        if hasattr(cells, '__iter__') and not hasattr(cells[0], '__iter__'):
            rows = [[c.value for c in cells]]
        else:
            rows = [[c.value for c in row] for row in cells]
    else:
        rows = [[c.value for c in row] for row in ws.iter_rows()]
    _out({"sheet": ws.title, "rows": rows})


def cmd_write(args):
    import openpyxl
    p = Path(args.file)
    wb = openpyxl.load_workbook(p) if p.exists() else openpyxl.Workbook()
    ws = wb[args.sheet] if args.sheet and args.sheet in wb.sheetnames else wb.active
    if args.sheet and args.sheet not in wb.sheetnames:
        ws = wb.create_sheet(args.sheet)
    data = json.loads(args.data)
    if args.range:
        start = ws[args.range.split(":")[0]]
        sr, sc = start.row, start.column
        if isinstance(data[0], list):
            for ri, row in enumerate(data):
                for ci, val in enumerate(row):
                    ws.cell(row=sr + ri, column=sc + ci, value=val)
        else:
            for ci, val in enumerate(data):
                ws.cell(row=sr, column=sc + ci, value=val)
    else:
        for ri, row in enumerate(data, 1):
            if isinstance(row, list):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)
            else:
                ws.cell(row=ri, column=1, value=row)
    wb.save(args.file)
    _out({"ok": True, "file": args.file, "sheet": ws.title})


def cmd_append(args):
    import openpyxl
    p = Path(args.file)
    wb = openpyxl.load_workbook(p) if p.exists() else openpyxl.Workbook()
    ws = wb[args.sheet] if args.sheet and args.sheet in wb.sheetnames else wb.active
    rows = json.loads(args.data)
    if not isinstance(rows[0], list):
        rows = [rows]
    for row in rows:
        ws.append(row)
    wb.save(args.file)
    _out({"ok": True, "new_max_row": ws.max_row})


def cmd_create(args):
    import openpyxl
    wb = openpyxl.Workbook()
    if args.sheet:
        wb.active.title = args.sheet
    wb.save(args.file)
    _out({"ok": True, "file": args.file, "sheets": wb.sheetnames})


def cmd_add_sheet(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.file)
    if args.sheet in wb.sheetnames:
        _out({"ok": False, "error": f"Sheet '{args.sheet}' already exists"})
        return
    wb.create_sheet(args.sheet)
    wb.save(args.file)
    _out({"ok": True, "sheets": wb.sheetnames})


def cmd_delete_sheet(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.file)
    if args.sheet not in wb.sheetnames:
        _out({"ok": False, "error": f"Sheet '{args.sheet}' not found"})
        return
    del wb[args.sheet]
    wb.save(args.file)
    _out({"ok": True, "sheets": wb.sheetnames})


def cmd_clear(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    for row in ws[args.range]:
        for cell in (row if hasattr(row, '__iter__') else [row]):
            cell.value = None
    wb.save(args.file)
    _out({"ok": True})


def cmd_delete_rows(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    ws.delete_rows(args.start, args.count)
    wb.save(args.file)
    _out({"ok": True, "new_max_row": ws.max_row})


def cmd_delete_cols(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.file)
    ws = wb[args.sheet] if args.sheet else wb.active
    ws.delete_cols(args.start, args.count)
    wb.save(args.file)
    _out({"ok": True, "new_max_col": ws.max_column})


def main():
    p = argparse.ArgumentParser(description="Excel CLI for Claude Code")
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_file(sp):
        sp.add_argument("--file", "-f", required=True, help="Path to .xlsx file")

    def add_sheet(sp):
        sp.add_argument("--sheet", "-s", default=None, help="Sheet name (default: active)")

    # list-sheets
    s = sub.add_parser("list-sheets", help="List sheet names")
    add_file(s); s.set_defaults(func=cmd_list_sheets)

    # info
    s = sub.add_parser("info", help="Sheet info: dims, max_row, max_col")
    add_file(s); add_sheet(s); s.set_defaults(func=cmd_info)

    # read
    s = sub.add_parser("read", help="Read cells, returns JSON rows")
    add_file(s); add_sheet(s)
    s.add_argument("--range", "-r", default=None, help="A1:B10 range (default: all used)")
    s.set_defaults(func=cmd_read)

    # write
    s = sub.add_parser("write", help="Write data to range. --data='[[row1],[row2]]'")
    add_file(s); add_sheet(s)
    s.add_argument("--range", "-r", default=None, help="Start cell or range (e.g. A1)")
    s.add_argument("--data", required=True, help="JSON: [[v1,v2],[v3,v4]] or [v1,v2]")
    s.set_defaults(func=cmd_write)

    # append
    s = sub.add_parser("append", help="Append rows to end of sheet")
    add_file(s); add_sheet(s)
    s.add_argument("--data", required=True, help="JSON: [[v1,v2]] or [v1,v2] for one row")
    s.set_defaults(func=cmd_append)

    # create
    s = sub.add_parser("create", help="Create new .xlsx file")
    add_file(s)
    s.add_argument("--sheet", "-s", default=None, help="First sheet name")
    s.set_defaults(func=cmd_create)

    # add-sheet
    s = sub.add_parser("add-sheet", help="Add worksheet")
    add_file(s)
    s.add_argument("--sheet", "-s", required=True)
    s.set_defaults(func=cmd_add_sheet)

    # delete-sheet
    s = sub.add_parser("delete-sheet", help="Delete worksheet")
    add_file(s)
    s.add_argument("--sheet", "-s", required=True)
    s.set_defaults(func=cmd_delete_sheet)

    # clear
    s = sub.add_parser("clear", help="Clear cell values in range")
    add_file(s); add_sheet(s)
    s.add_argument("--range", "-r", required=True)
    s.set_defaults(func=cmd_clear)

    # delete-rows
    s = sub.add_parser("delete-rows", help="Delete rows")
    add_file(s); add_sheet(s)
    s.add_argument("--start", type=int, required=True)
    s.add_argument("--count", type=int, default=1)
    s.set_defaults(func=cmd_delete_rows)

    # delete-cols
    s = sub.add_parser("delete-cols", help="Delete columns")
    add_file(s); add_sheet(s)
    s.add_argument("--start", type=int, required=True)
    s.add_argument("--count", type=int, default=1)
    s.set_defaults(func=cmd_delete_cols)

    args = p.parse_args()
    try:
        args.func(args)
    except Exception as e:
        _out({"ok": False, "error": str(e)})
        sys.exit(1)


if __name__ == "__main__":
    main()
